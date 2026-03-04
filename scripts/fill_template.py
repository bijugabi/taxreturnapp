#!/usr/bin/env python3
"""
fill_template.py
Preenche o template Excel oficial (MTD) com as transações extraídas pela IA.

Estrutura real do template (inspecionada):
  Sales:
    - Linha 1: "SALES" (título)
    - Linha 4: cabeçalhos Date | Inv # | Narrative | £ (colunas B=2, C=3, D=4, E=5)
    - Linhas 5–14: dados (máx. 10 entradas)
    - Linha 15, col E: =SUM(E5:E14)  ← fórmula fixa, não tocar
    - Linha 19: resumo de lucro (não tocar)

  Purchases:
    - Linha 1: "PURCHASES" (título)
    - Linha 4: cabeçalhos Date | Inv # | Narrative | £
    - Linhas 5–8: dados (máx. 4 entradas)
    - Linha 9, col E: =SUM(E5:E8)  ← fórmula fixa, não tocar

  ITSEFile: calculada automaticamente — NÃO TOCAR.

Uso:
  python3 fill_template.py '<json_string>' '<output_path>' '<template_path>'
"""

import sys
import json
import re
from datetime import datetime
from openpyxl import load_workbook

# ─── Estrutura base do template ───────────────────────────────────────────────
SALES_START_ROW = 5      # primeira linha de dados em 'Sales'
SALES_SUM_ROW = 15       # linha onde fica a fórmula de soma das vendas

PURCHASES_START_ROW = 5  # primeira linha de dados em 'Purchases'
PURCHASES_SUM_ROW = 9    # linha onde fica a fórmula de soma das compras

DATA_COL_START = 2       # coluna B

MONTH_ABBR = [
    None, "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC",
]


def parse_date(date_str: str):
    """Converte data ISO (YYYY-MM-DD) para objeto datetime (o Excel formata automaticamente)."""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        return date_str


def clear_data_rows(sheet, start_row: int, end_row: int, start_col: int = 2, num_cols: int = 4):
    """Limpa as linhas de dados antes de preencher (remove dados de exemplo do template)."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, start_col + num_cols):
            sheet.cell(row=row, column=col).value = None


def fill_sheet(sheet, transactions: list, start_row: int, end_row: int):
    """
    Preenche as linhas de dados de uma aba.
    Colunas B=Date, C=Inv #, D=Narrative, E=£
    """
    for i, txn in enumerate(transactions):
        row = start_row + i
        date_cell = sheet.cell(row=row, column=DATA_COL_START)
        date_cell.value = parse_date(txn.get("date", ""))
        date_cell.number_format = "DD/MM/YYYY"
        sheet.cell(row=row, column=DATA_COL_START + 1).value = str(txn.get("invoice", "N/A"))
        sheet.cell(row=row, column=DATA_COL_START + 2).value = str(txn.get("narrative", ""))
        amount = txn.get("amount", 0)
        try:
            sheet.cell(row=row, column=DATA_COL_START + 3).value = float(amount)
        except (TypeError, ValueError):
            sheet.cell(row=row, column=DATA_COL_START + 3).value = 0.0


def main():
    if len(sys.argv) < 4:
        print(
            "Uso: python3 fill_template.py '<json>' '<output_path>' '<template_path>'",
            file=sys.stderr
        )
        sys.exit(1)

    json_str      = sys.argv[1]
    output_path   = sys.argv[2]
    template_path = sys.argv[3]
    quarter       = sys.argv[4] if len(sys.argv) > 4 else "Q1"
    year         = sys.argv[5] if len(sys.argv) > 5 else ""

    try:
        data = json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"Erro ao parsear JSON: {e}", file=sys.stderr)
        sys.exit(1)

    raw_sales     = data.get("sales", []) or []
    raw_purchases = data.get("purchases", []) or []

    quarter_key = str(quarter).upper()

    def month_from_date(date_str: str) -> int:
        """Extrai o mês (1-12) da data em formato YYYY-MM-DD."""
        if not date_str or not isinstance(date_str, str):
            return 1
        try:
            dt = datetime.strptime(date_str.strip()[:10], "%Y-%m-%d")
            return dt.month
        except Exception:
            return 1

    def build_entries(transactions, quarter_prefix: str):
        result = []
        for idx, txn in enumerate(transactions, start=1):
            date_str = txn.get("date", "")
            month_num = month_from_date(date_str)
            month_abbr = MONTH_ABBR[month_num] if 1 <= month_num <= 12 else "N/A"
            inv = f"{quarter_prefix}{idx:03d}"
            result.append(
                {
                    "date": date_str,
                    "invoice": inv,
                    "narrative": f"Invoice for services in {month_abbr}",
                    "amount": txn.get("amount", 0),
                }
            )
        return result

    sales = build_entries(raw_sales, quarter_key)
    purchases = build_entries(raw_purchases, quarter_key)

    print(f"  Sales a preencher: {len(sales)}")
    print(f"  Purchases a preencher: {len(purchases)}")

    try:
        wb = load_workbook(template_path)
    except Exception as e:
        print(f"Erro ao abrir template: {e}", file=sys.stderr)
        sys.exit(1)

    sales_total_row = SALES_SUM_ROW
    purchases_total_row = PURCHASES_SUM_ROW

    # ── Preenche aba Sales ────────────────────────────────────────────────────
    if "Sales" in wb.sheetnames:
        ws_sales = wb["Sales"]
        sales_end_row = SALES_START_ROW + max(len(sales), 1) - 1

        if sales_end_row >= SALES_SUM_ROW:
            extra_rows = sales_end_row - (SALES_SUM_ROW - 1)
            ws_sales.insert_rows(SALES_SUM_ROW, amount=extra_rows)
            sales_total_row = SALES_SUM_ROW + extra_rows
        else:
            sales_total_row = SALES_SUM_ROW

        clear_data_rows(ws_sales, SALES_START_ROW, sales_end_row)
        fill_sheet(ws_sales, sales, SALES_START_ROW, sales_end_row)
        ws_sales.cell(row=sales_total_row, column=5).value = (
            f"=SUM(E{SALES_START_ROW}:E{sales_end_row})"
        )
    else:
        print("AVISO: aba 'Sales' nao encontrada no template.", file=sys.stderr)

    # ── Preenche aba Purchases ────────────────────────────────────────────────
    if "Purchases" in wb.sheetnames:
        ws_purchases = wb["Purchases"]
        purchases_end_row = PURCHASES_START_ROW + max(len(purchases), 1) - 1

        if purchases_end_row >= PURCHASES_SUM_ROW:
            extra_rows = purchases_end_row - (PURCHASES_SUM_ROW - 1)
            ws_purchases.insert_rows(PURCHASES_SUM_ROW, amount=extra_rows)
            purchases_total_row = PURCHASES_SUM_ROW + extra_rows
        else:
            purchases_total_row = PURCHASES_SUM_ROW

        clear_data_rows(ws_purchases, PURCHASES_START_ROW, purchases_end_row)
        fill_sheet(ws_purchases, purchases, PURCHASES_START_ROW, purchases_end_row)
        ws_purchases.cell(row=purchases_total_row, column=5).value = (
            f"=SUM(E{PURCHASES_START_ROW}:E{purchases_end_row})"
        )
    else:
        print("AVISO: aba 'Purchases' nao encontrada no template.", file=sys.stderr)

    # ── Atualiza referências ao total em todas as abas ─────────────────────────
    # Fórmulas que usam Sales!E15 ou Purchases!E9 passam a apontar para a linha real do total.
    # Profit for Q2 na aba Sales: troca =SUM(E18:E21) ou células fixas pelo total real de Sales e Purchases.
    profit_formula_new = f"=E{sales_total_row}-'Purchases'!E{purchases_total_row}"
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    formula = formula.replace("'Sales'!E15", f"'Sales'!E{sales_total_row}")
                    formula = formula.replace("Sales!E15", f"Sales!E{sales_total_row}")
                    formula = formula.replace("'Purchases'!E9", f"'Purchases'!E{purchases_total_row}")
                    formula = formula.replace("Purchases!E9", f"Purchases!E{purchases_total_row}")
                    if sheet.title == "Sales":
                        formula = re.sub(r"\bE15\b", f"E{sales_total_row}", formula)
                        formula = formula.replace("=SUM(E18:E21)", profit_formula_new)
                        formula = formula.replace("=E18-E21", profit_formula_new)
                        formula = formula.replace("=E18-'Purchases'!E9", profit_formula_new)
                    elif sheet.title == "Purchases":
                        formula = re.sub(r"\bE9\b", f"E{purchases_total_row}", formula)
                    cell.value = formula

    try:
        wb.save(output_path)
        print(f"OK: arquivo salvo em {output_path}")
    except Exception as e:
        print(f"Erro ao salvar arquivo: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
